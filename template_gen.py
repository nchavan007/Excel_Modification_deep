from openpyxl import load_workbook

# load main  file
main_workbook = load_workbook(filename="sample.xlsx")

# load template  file
template1 = load_workbook(filename="template1.xlsx")
template2 = load_workbook(filename="template2.xlsx")
template3 = load_workbook(filename="template3.xlsx")
template4 = load_workbook(filename="template4.xlsx")

temp1_sheet = template1.sheetnames
temp2_sheet = template2.sheetnames
temp3_sheet = template3.sheetnames
temp4_sheet = template4.sheetnames

# open workbook
main_sheet = main_workbook.active

tuple_list = [('B',1), ('C',2), ( 'D', 3)]

for row in main_sheet.iter_rows(min_row=2, values_only=True):
   for i, j in tuple_list:
       for k in temp1_sheet:
           template1[k][f'{i}2'] = row[j]

       for k in temp2_sheet:
           template2[k][f'{i}2'] = row[j]

       for k in temp3_sheet:
           template3[k][f'{i}2'] = row[j]

       for k in temp4_sheet:
           template4[k][f'{i}2'] = row[j]

   template1.save(filename=f'template1_{row[0]}.xlsx')
   template2.save(filename=f'template2_{row[0]}.xlsx')
   template3.save(filename=f'template3_{row[0]}.xlsx')
   template4.save(filename=f'template4_{row[0]}.xlsx')
